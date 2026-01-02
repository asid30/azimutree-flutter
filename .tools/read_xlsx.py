from openpyxl import load_workbook
import json
from datetime import datetime

path = "assets/templates/template_azimutree_import_data_v2.xlsx"
wb = load_workbook(filename=path, read_only=True, data_only=True)
output = {}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        # normalize values
        norm = []
        for v in row:
            if isinstance(v, datetime):
                norm.append(v.isoformat())
            else:
                norm.append(v)
        rows.append(norm)
    output[sheet] = rows
print(json.dumps(output, ensure_ascii=False, indent=2))
